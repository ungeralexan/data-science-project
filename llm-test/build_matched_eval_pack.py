"""
Build a minimal, auditable evaluation pack AFTER a single LLM run.

Inputs (from tests/local_eval):
- llm_events_<timestamp>.json
- all_emails_raw_<timestamp>.txt   (recommended) OR all_emails_filtered_<timestamp>.txt

Outputs (to tests/local_eval):
- matched_emails_<timestamp>.txt          (only emails referenced by events)
- matched_emails_index_<timestamp>.json   (subject/from lookup)
- matched_eval_<timestamp>.xlsx           (Events + MatchedEmails)
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Set
import argparse
import json
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


LOCAL_EVAL_DIR = Path("tests/local_eval")

EMAIL_BLOCK_RE = re.compile(
    r"--------------- EMAIL:\s*(\d+)\s*Start ---------------(.*?)--------------- EMAIL:\s*\1\s*End ---------------",
    re.DOTALL,
)


def split_email_blocks(all_emails_text: str) -> Dict[int, Dict[str, Any]]:
    """
    Return dict keyed by Source_Email_Index:
      { idx: {Source_Email_Index, Subject, From, Raw_Block} }
    """
    out: Dict[int, Dict[str, Any]] = {}
    for m in EMAIL_BLOCK_RE.finditer(all_emails_text):
        idx = int(m.group(1))
        raw_block = m.group(0)
        inner = m.group(2)

        subj = ""
        frm = ""
        subj_m = re.search(r"Subject:\s*(.*)", inner)
        frm_m = re.search(r"From:\s*(.*)", inner)
        if subj_m:
            subj = subj_m.group(1).strip()
        if frm_m:
            frm = frm_m.group(1).strip()

        out[idx] = {
            "Source_Email_Index": idx,
            "Subject": subj,
            "From": frm,
            "Raw_Block": raw_block,
        }
    return out


def save_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def save_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def save_excel(path: Path, events: List[Dict[str, Any]], matched_emails: List[Dict[str, Any]]) -> None:
    wb = Workbook()

    # Sheet 1: Events (your main deliverable)
    ws1 = wb.active
    ws1.title = "Events"

    headers_events = [
        "event_index",
        "Source_Email_Index",
        "Title", "Start_Date", "End_Date", "Start_Time", "End_Time",
        "Location", "Organizer", "Speaker",
        "Registration_Needed", "URL",
        "Event_Type", "Main_Event_Temp_Key",
        "Evidence",
        "Description",
    ]
    ws1.append(headers_events)

    for i, ev in enumerate(events, start=1):
        ws1.append([
            i,
            ev.get("Source_Email_Index", ""),
            ev.get("Title", ""),
            ev.get("Start_Date", ""),
            ev.get("End_Date", ""),
            ev.get("Start_Time", ""),
            ev.get("End_Time", ""),
            ev.get("Location", ""),
            ev.get("Organizer", ""),
            ev.get("Speaker", ""),
            ev.get("Registration_Needed", ""),
            ev.get("URL", ""),
            ev.get("Event_Type", ""),
            ev.get("Main_Event_Temp_Key", ""),
            ev.get("Evidence", ""),
            ev.get("Description", ""),
        ])

    # Sheet 2: MatchedEmails (only those referenced by events)
    ws2 = wb.create_sheet("MatchedEmails")
    headers_emails = ["Source_Email_Index", "Subject", "From", "Preview"]
    ws2.append(headers_emails)

    for em in matched_emails:
        raw = em.get("Raw_Block", "")
        preview = re.sub(r"\s+", " ", raw)[:350]
        ws2.append([
            em.get("Source_Email_Index", ""),
            em.get("Subject", ""),
            em.get("From", ""),
            preview,
        ])

    # Column widths (simple, readable)
    for ws in [ws1, ws2]:
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 24

    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--events-json", required=True, help="Path to llm_events_<ts>.json")
    parser.add_argument("--emails-txt", required=True, help="Path to all_emails_raw_<ts>.txt (or filtered)")
    parser.add_argument("--out-prefix", required=True, help="Prefix for output files, e.g. matched_20251226_120000")
    args = parser.parse_args()

    events_path = Path(args.events_json)
    emails_path = Path(args.emails_txt)

    events = json.loads(events_path.read_text(encoding="utf-8"))
    all_emails_text = emails_path.read_text(encoding="utf-8")

    if not isinstance(events, list):
        raise ValueError("events JSON is not a list")

    # Which email indices do we need?
    needed: Set[int] = set()
    for ev in events:
        try:
            needed.add(int(ev.get("Source_Email_Index")))
        except Exception:
            pass

    # Parse all email blocks
    blocks = split_email_blocks(all_emails_text)

    # Extract matched emails, keep sorted by index
    matched = []
    missing = []
    for idx in sorted(needed):
        if idx in blocks:
            matched.append(blocks[idx])
        else:
            missing.append(idx)

    # Outputs
    out_prefix = args.out_prefix
    out_txt = LOCAL_EVAL_DIR / f"{out_prefix}_matched_emails.txt"
    out_idx = LOCAL_EVAL_DIR / f"{out_prefix}_matched_emails_index.json"
    out_xlsx = LOCAL_EVAL_DIR / f"{out_prefix}_matched_eval.xlsx"
    out_report = LOCAL_EVAL_DIR / f"{out_prefix}_report.json"

    # Save matched emails text
    matched_text = "\n\n".join([m["Raw_Block"] for m in matched])
    save_text(out_txt, matched_text)

    # Save index JSON
    index_json = [
        {"Source_Email_Index": m["Source_Email_Index"], "Subject": m["Subject"], "From": m["From"]}
        for m in matched
    ]
    save_json(out_idx, index_json)

    # Save Excel
    save_excel(out_xlsx, events, matched)

    # Save a small report (useful for you / prof explanation)
    report = {
        "events_count": len(events),
        "unique_source_email_indices": len(needed),
        "matched_email_blocks_found": len(matched),
        "missing_email_indices": missing,
        "events_json": str(events_path),
        "emails_txt": str(emails_path),
        "outputs": {
            "matched_emails_txt": str(out_txt),
            "matched_emails_index_json": str(out_idx),
            "matched_eval_xlsx": str(out_xlsx),
        },
    }
    save_json(out_report, report)

    print("[matched-pack] Done.")
    print(f"[matched-pack] Events: {len(events)}")
    print(f"[matched-pack] Unique Source_Email_Index: {len(needed)}")
    print(f"[matched-pack] Matched email blocks found: {len(matched)}")
    if missing:
        print(f"[matched-pack] WARNING: missing email indices: {missing}")
    print(f"[matched-pack] Wrote: {out_xlsx}")


if __name__ == "__main__":
    main()
