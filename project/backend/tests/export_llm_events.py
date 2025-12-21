"""
Export LLM-extracted events directly to JSON (always) and optionally Excel (.xlsx).

Usage (from project/backend):

    python -m tests.export_llm_events_xlsx --limit 20
    python -m tests.export_llm_events_xlsx --limit 50 --no-excel
"""

from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import argparse
import json

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from config import EMAIL_TEMP_DIR, EMAIL_PIPELINE_DEFAULT_LIMIT  # type: ignore
from services.email_downloader import download_latest_emails  # type: ignore
from services.event_recognizer import extract_event_info_with_llm  # type: ignore



# Strictly local, gitignored (contains private content)
LOCAL_EVAL_DIR = Path("local_eval")



def load_all_emails_text(limit: int) -> str:
    """Download and load the combined all_emails.txt."""
    print(f"[eval] Downloading latest {limit} emails...")
    download_latest_emails(limit=limit)

    all_emails_path = Path(EMAIL_TEMP_DIR) / "all_emails.txt"
    if not all_emails_path.exists():
        raise FileNotFoundError(f"[eval] Combined email file not found at {all_emails_path}")

    print(f"[eval] Using: {all_emails_path}")
    return all_emails_path.read_text(encoding="utf-8")


def save_events_to_json(events: List[Dict[str, Any]], timestamp: str) -> Path:
    """Save events into a local JSON file (gitignored)."""
    LOCAL_EVAL_DIR.mkdir(parents=True, exist_ok=True)
    json_path = LOCAL_EVAL_DIR / f"llm_events_{timestamp}.json"
    json_path.write_text(
        json.dumps(events, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"[eval] JSON file saved: {json_path}")
    return json_path


def save_events_to_excel(events: List[Dict[str, Any]], timestamp: str) -> Path:
    """Save events into an Excel file for manual inspection."""
    LOCAL_EVAL_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = LOCAL_EVAL_DIR / f"llm_events_{timestamp}.xlsx"


    wb = Workbook()
    ws = wb.active
    ws.title = "LLM Events"

    headers = [
        "event_index",
        "Title",
        "Start_Date",
        "End_Date",
        "Start_Time",
        "End_Time",
        "Location",
        "Organizer",
        "Speaker",
        "Description",
        "Image_Key",
    ]
    ws.append(headers)

    for idx, ev in enumerate(events, start=1):
        ws.append([
            idx,
            ev.get("Title", ""),
            ev.get("Start_Date", ""),
            ev.get("End_Date", ""),
            ev.get("Start_Time", ""),
            ev.get("End_Time", ""),
            ev.get("Location", ""),
            ev.get("Organizer", ""),
            ev.get("Speaker", ""),
            ev.get("Description", ""),
            ev.get("Image_Key", ""),
        ])

    for col_idx, column_title in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = min(max(len(column_title), 20), 50)

    wb.save(xlsx_path)
    print(f"[eval] Excel file saved: {xlsx_path}")
    return xlsx_path


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--limit",
        type=int,
        default=EMAIL_PIPELINE_DEFAULT_LIMIT,
        help="How many latest emails to fetch",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="If set, do not create an Excel file (JSON export still happens).",
    )
    args = parser.parse_args()

    all_emails = load_all_emails_text(limit=args.limit)

    print("[eval] Running LLM extraction...")
    events = extract_event_info_with_llm(all_emails)

    if not isinstance(events, list):
        print("[eval] ERROR: LLM returned non-list structure")
        return

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Always save JSON locally (gitignored)
    save_events_to_json(events, timestamp=timestamp)

    # Optionally save Excel
    if args.no_excel:
        print("[eval] Skipping Excel export (--no-excel).")
    else:
        save_events_to_excel(events, timestamp=timestamp)


if __name__ == "__main__":
    main()
