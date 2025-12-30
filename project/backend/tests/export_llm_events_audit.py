"""
Export LLM-extracted events with auditability:
- Prefilter emails to reduce trash (no extra LLM calls)
- Force each event to include Source_Email_Index + Evidence snippet
- Export JSON + Excel with two sheets (Events + EmailsUsed)

Run (from project/backend):
  python -m tests.export_llm_events_audit --limit 150
"""

from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Tuple
from datetime import datetime
import argparse
import json
import re

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from google import genai

from config import EMAIL_TEMP_DIR, EMAIL_PIPELINE_DEFAULT_LIMIT, RECOGNITION_LLM_MODEL  # pylint: disable=import-error
from services.email_downloader.email_downloader import download_latest_emails  # pylint: disable=import-error


# Strictly local, gitignored (contains private content)
LOCAL_EVAL_DIR = Path("tests/local_eval")


# ----------------------------
# 1) Parsing email blocks
# ----------------------------
EMAIL_BLOCK_RE = re.compile(
    r"--------------- EMAIL:\s*(\d+)\s*Start ---------------(.*?)--------------- EMAIL:\s*\1\s*End ---------------",
    re.DOTALL,
)

def split_email_blocks(all_emails_text: str) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    for m in EMAIL_BLOCK_RE.finditer(all_emails_text):
        idx = int(m.group(1))
        raw = m.group(0)
        inner = m.group(2)

        # Extract subject/from quickly (best-effort)
        subj = ""
        frm = ""
        subj_m = re.search(r"Subject:\s*(.*)", inner)
        frm_m = re.search(r"From:\s*(.*)", inner)
        if subj_m:
            subj = subj_m.group(1).strip()
        if frm_m:
            frm = frm_m.group(1).strip()

        blocks.append({
            "Source_Email_Index": idx,
            "Subject": subj,
            "From": frm,
            "Raw_Block": raw,
        })
    return blocks


# ----------------------------
# 2) Heuristic prefilter
# ----------------------------
EVENT_CUE_PATTERNS = [
    r"\bdatum\b", r"\buhrzeit\b", r"\bort\b",                # DE cues
    r"\bdate\b", r"\btime\b", r"\blocation\b",              # EN cues
    r"\bworkshop\b", r"\blecture\b", r"\btalk\b", r"\bseminar\b",
    r"\bregistration\b", r"\banmeldung\b",
    r"\bzoom\b", r"\bonline\b", r"\bhybrid\b",
    r"\b\d{1,2}\.\d{1,2}\.\d{4}\b",                          # 12.01.2025
    r"\b\d{4}-\d{2}-\d{2}\b",                                # 2025-01-12
    r"\b\d{1,2}/\d{1,2}/\d{4}\b",                            # 01/12/2025
    r"\b\d{1,2}:\d{2}\b",                                    # 14:30
    r"\b(c\.t\.|s\.t\.)\b",
]

EVENT_CUE_RE = re.compile("|".join(EVENT_CUE_PATTERNS), re.IGNORECASE)

def likely_contains_event(email_block: str) -> bool:
    # Simple but effective: if no event cue appears, skip it
    return bool(EVENT_CUE_RE.search(email_block))


def prefilter_blocks(blocks: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    kept: List[Dict[str, Any]] = []
    for b in blocks:
        if likely_contains_event(b["Raw_Block"]):
            kept.append(b)
    return kept


# ----------------------------
# 3) LLM extraction (tests-only)
# ----------------------------
EVENT_SCHEMA_AUDIT = {
    "type": "OBJECT",
    "properties": {
        "Title": {"type": "STRING", "nullable": True},
        "Start_Date": {"type": "STRING", "nullable": True},
        "End_Date": {"type": "STRING", "nullable": True},
        "Start_Time": {"type": "STRING", "nullable": True},
        "End_Time": {"type": "STRING", "nullable": True},
        "Description": {"type": "STRING", "nullable": True},
        "Location": {"type": "STRING", "nullable": True},
        "Street": {"type": "STRING", "nullable": True},
        "House_Number": {"type": "STRING", "nullable": True},
        "Zip_Code": {"type": "STRING", "nullable": True},
        "City": {"type": "STRING", "nullable": True},
        "Country": {"type": "STRING", "nullable": True},
        "Room": {"type": "STRING", "nullable": True},
        "Floor": {"type": "STRING", "nullable": True},
        "Speaker": {"type": "STRING", "nullable": True},
        "Organizer": {"type": "STRING", "nullable": True},
        "Registration_Needed": {"type": "BOOLEAN", "nullable": True},
        "URL": {"type": "STRING", "nullable": True},
        "Image_Key": {"type": "STRING", "nullable": True},
        "Event_Type": {"type": "STRING", "nullable": False},
        "Main_Event_Temp_Key": {"type": "STRING", "nullable": False},

        # Audit fields (critical)
        "Source_Email_Index": {"type": "INTEGER", "nullable": False},
        "Evidence": {"type": "STRING", "nullable": True},
    },
    "required": [
        "Title","Start_Date","End_Date","Start_Time","End_Time","Description","Location",
        "Street","House_Number","Zip_Code","City","Country","Room","Floor","Speaker","Organizer",
        "Registration_Needed","URL","Image_Key","Event_Type","Main_Event_Temp_Key",
        "Source_Email_Index","Evidence"
    ],
}

SCHEMA_MULTI_AUDIT = {"type": "ARRAY", "items": EVENT_SCHEMA_AUDIT}


def extract_events_with_audit_llm(filtered_email_text: str) -> List[Dict[str, Any]]:
    # Load Gemini API key
    with open("secrets.json", "r", encoding="utf-8") as f:
        secrets = json.load(f)

    client = genai.Client(api_key=secrets["GEMINI_API_KEY"])

    system_instruction = """
You extract structured university events from concatenated email blocks.

CRITICAL AUDIT REQUIREMENTS:
1) For every event you return, you MUST set Source_Email_Index to the EMAIL index number shown in:
   "--------------- EMAIL: X Start ---------------"
2) You MUST provide an Evidence string: a short quote (max ~250 chars) copied from the same email block
   (body or URL content) that supports the date/time/location/title decision.
   Evidence should include at least one concrete detail (date OR time OR location).

GENERAL RULES:
- If an email contains no event, produce no event for it.
- Output dates as MM/DD/YYYY and times as HH:MM AM/PM when possible, else null.
- Event_Type must be "main_event" or "sub_event".
- Sub events must share Main_Event_Temp_Key with their parent main_event.

Return ONLY a JSON ARRAY of event objects following the schema. No extra text.
""".strip()

    user_prompt = f"""
Extract events from the following concatenated emails:

{filtered_email_text}
""".strip()

    resp = client.models.generate_content(
        model=RECOGNITION_LLM_MODEL,
        contents=f"{system_instruction}\n\n{user_prompt}",
        config={
            "response_mime_type": "application/json",
            "response_schema": SCHEMA_MULTI_AUDIT,
        },
    )

    return json.loads(resp.text)


# ----------------------------
# 4) Export utilities
# ----------------------------
def save_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def save_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def save_excel(path: Path, events: List[Dict[str, Any]], emails_used: List[Dict[str, Any]]) -> None:
    wb = Workbook()

    # Sheet 1: Events
    ws1 = wb.active
    ws1.title = "Events"

    headers_events = [
        "event_index",
        "Source_Email_Index",
        "Title","Start_Date","End_Date","Start_Time","End_Time",
        "Location","Organizer","Speaker","Registration_Needed","URL",
        "Event_Type","Main_Event_Temp_Key",
        "Evidence",
        "Description",
    ]
    ws1.append(headers_events)

    for i, ev in enumerate(events, start=1):
        ws1.append([
            i,
            ev.get("Source_Email_Index", ""),
            ev.get("Title",""),
            ev.get("Start_Date",""),
            ev.get("End_Date",""),
            ev.get("Start_Time",""),
            ev.get("End_Time",""),
            ev.get("Location",""),
            ev.get("Organizer",""),
            ev.get("Speaker",""),
            ev.get("Registration_Needed",""),
            ev.get("URL",""),
            ev.get("Event_Type",""),
            ev.get("Main_Event_Temp_Key",""),
            ev.get("Evidence",""),
            ev.get("Description",""),
        ])

    # Sheet 2: EmailsUsed (for lookup)
    ws2 = wb.create_sheet("EmailsUsed")
    headers_emails = ["Source_Email_Index", "Subject", "From", "Preview"]
    ws2.append(headers_emails)

    for e in emails_used:
        raw = e.get("Raw_Block","")
        preview = re.sub(r"\s+", " ", raw)[:300]
        ws2.append([
            e.get("Source_Email_Index",""),
            e.get("Subject",""),
            e.get("From",""),
            preview,
        ])

    # Column widths
    for ws in [ws1, ws2]:
        for col_idx, title in enumerate(ws[1], start=1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 22

    wb.save(path)


# ----------------------------
# 5) Main
# ----------------------------
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--limit", type=int, default=EMAIL_PIPELINE_DEFAULT_LIMIT)
    parser.add_argument("--no-excel", action="store_true")
    args = parser.parse_args()

    print(f"[audit] Downloading latest {args.limit} emails...")
    download_latest_emails(limit=args.limit)

    all_emails_path = Path(EMAIL_TEMP_DIR) / "all_emails.txt"
    if not all_emails_path.exists():
        raise FileNotFoundError(f"[audit] Missing: {all_emails_path}")

    all_text = all_emails_path.read_text(encoding="utf-8")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Parse + prefilter
    blocks = split_email_blocks(all_text)
    kept = prefilter_blocks(blocks)

    print(f"[audit] Parsed {len(blocks)} email blocks, kept {len(kept)} after prefilter.")

    # Build filtered concatenated text (only kept blocks)
    filtered_text = "\n\n".join([b["Raw_Block"] for b in kept])

    # Save artifacts
    LOCAL_EVAL_DIR.mkdir(parents=True, exist_ok=True)
    save_text(LOCAL_EVAL_DIR / f"all_emails_raw_{timestamp}.txt", all_text)
    save_text(LOCAL_EVAL_DIR / f"all_emails_filtered_{timestamp}.txt", filtered_text)

    # Save index mapping for traceability
    email_index = [
        {
            "Source_Email_Index": b["Source_Email_Index"],
            "Subject": b["Subject"],
            "From": b["From"],
        }
        for b in kept
    ]
    save_json(LOCAL_EVAL_DIR / f"email_index_{timestamp}.json", email_index)

    # ONE LLM CALL
    print("[audit] Running ONE LLM extraction call on filtered emails...")
    events = extract_events_with_audit_llm(filtered_text)
    if not isinstance(events, list):
        raise ValueError("[audit] LLM output is not a list")

    save_json(LOCAL_EVAL_DIR / f"llm_events_{timestamp}.json", events)

    if args.no_excel:
        print("[audit] Skipping Excel (--no-excel).")
    else:
        xlsx_path = LOCAL_EVAL_DIR / f"llm_events_{timestamp}.xlsx"
        save_excel(xlsx_path, events, kept)
        print(f"[audit] Excel saved: {xlsx_path}")

    print(f"[audit] Done. Events extracted: {len(events)}")


if __name__ == "__main__":
    main()
