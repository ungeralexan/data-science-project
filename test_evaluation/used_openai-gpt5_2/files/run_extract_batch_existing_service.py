from __future__ import annotations

import argparse
import json
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# IMPORTANT: uses your existing services extractor (do not modify services)
from backend.services.event_recognizer import extract_event_info_with_llm  # pylint: disable=import-error


DATASET_DIR = Path("backend/tests/evaluation/data/raw_emails_150")
PRED_ROOT = Path("backend/tests/evaluation/predictions")

# The “must-check” fields you agreed on (export subset)
FILTER_FIELDS = [
    "Title",
    "Start_Date", "Start_Time", "End_Date", "End_Time",
    "Location",
    "Registration_Needed",
    "URL", "Registration_URL", "Meeting_URL",
    "Speaker", "Organizer",
    "Event_Type", "Main_Event_Temp_Key",
]

def load_index() -> Dict[str, Any]:
    idx_path = DATASET_DIR / "index.json"
    if not idx_path.exists():
        raise FileNotFoundError(f"Missing dataset index: {idx_path}")
    return json.loads(idx_path.read_text(encoding="utf-8"))

def load_email_json(email_id: str) -> Dict[str, Any]:
    p = DATASET_DIR / "emails_json" / f"{email_id}.json"
    if not p.exists():
        raise FileNotFoundError(f"Missing per-email JSON: {p}")
    return json.loads(p.read_text(encoding="utf-8"))

def build_batch_input_text(batch_name: str, email_ids: List[str]) -> str:
    """
    Build the exact concatenated format expected by the existing system prompt:
      --------------- EMAIL: X Start ---------------
      Subject: ...
      From: ...
      Body: ...
      --------------- EMAIL: X End ---------------
    """
    blocks: List[str] = []
    for i, eid in enumerate(email_ids, start=1):
        em = load_email_json(eid)
        subject = em.get("subject", "") or ""
        sender = em.get("from", "") or ""
        body = em.get("body_text", "") or ""

        block = (
            f"--------------- EMAIL: {i} Start ---------------\n"
            f"Subject: {subject}\n"
            f"From: {sender}\n\n"
            f"Body: {body}\n"
            f"--------------- EMAIL: {i} End ---------------\n"
        )
        blocks.append(block)

    header = f"# Batch: {batch_name}\n# Emails: {len(email_ids)}\n\n"
    return header + "\n\n".join(blocks)

def filter_events(events: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Keep only the critical fields for human review + evaluation.
    We still save the full output separately.
    """
    filtered: List[Dict[str, Any]] = []
    for ev in events:
        row = {k: ev.get(k) for k in FILTER_FIELDS}
        filtered.append(row)
    return filtered

def save_json(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")

def save_excel(path: Path, filtered_events: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"

    ws.append(FILTER_FIELDS)

    for ev in filtered_events:
        ws.append([ev.get(k, None) for k in FILTER_FIELDS])

    # Column widths
    for col_idx in range(1, len(FILTER_FIELDS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)

def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--batch", type=str, default="batch_01", help="e.g., batch_01 ... batch_05")
    args = parser.parse_args()

    idx = load_index()
    batches = idx.get("batches", {})
    if args.batch not in batches:
        raise KeyError(f"Batch not found in index.json: {args.batch}")

    email_ids: List[str] = batches[args.batch]
    batch_text = build_batch_input_text(args.batch, email_ids)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = PRED_ROOT / f"{args.batch}_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    # Save exact input for reproducibility
    (out_dir / "input_sent_to_llm.txt").write_text(batch_text, encoding="utf-8")

    # ONE LLM CALL (existing services function)
    events = extract_event_info_with_llm(batch_text)
    if not isinstance(events, list):
        raise ValueError("LLM output is not a list; cannot proceed safely.")

    # Save full output (audit/debug)
    save_json(out_dir / "events_full.json", events)

    # Save filtered output (what you will share and inspect)
    filtered = filter_events(events)
    save_json(out_dir / "events_filtered.json", filtered)
    save_excel(out_dir / "events_filtered.xlsx", filtered)

    print(f"[extract_batch] Batch: {args.batch}")
    print(f"[extract_batch] Events extracted: {len(events)}")
    print(f"[extract_batch] Saved folder: {out_dir}")

if __name__ == "__main__":
    main()
