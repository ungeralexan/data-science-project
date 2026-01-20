from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


PRED_ROOT = Path("backend/tests/evaluation/predictions")
OUT_ROOT = Path("backend/tests/evaluation/gold_selection")

# --- Heuristic cues (for picking borderline/negative examples without LLM calls) ---
EVENT_CUE_PATTERNS = [
    r"\bdatum\b", r"\buhrzeit\b", r"\bort\b",
    r"\bdate\b", r"\btime\b", r"\blocation\b",
    r"\bworkshop\b", r"\blecture\b", r"\btalk\b", r"\bseminar\b",
    r"\bregistration\b", r"\banmeldung\b",
    r"\bzoom\b", r"\bonline\b", r"\bhybrid\b",
    r"\b\d{1,2}\.\d{1,2}\.\d{4}\b",
    r"\b\d{4}-\d{2}-\d{2}\b",
    r"\b\d{1,2}/\d{1,2}/\d{4}\b",
    r"\b\d{1,2}:\d{2}\b",
    r"\b(c\.t\.|s\.t\.)\b",
]
EVENT_CUE_RE = re.compile("|".join(EVENT_CUE_PATTERNS), re.IGNORECASE)


EMAIL_BLOCK_RE = re.compile(
    r"--------------- EMAIL:\s*(\d+)\s*Start ---------------(.*?)--------------- EMAIL:\s*\1\s*End ---------------",
    re.DOTALL,
)

URL_RE = re.compile(r"https?://\S+", re.IGNORECASE)


@dataclass
class EmailBlock:
    idx: int
    raw_block: str


def parse_email_blocks(batch_input_text: str) -> List[EmailBlock]:
    blocks: List[EmailBlock] = []
    for m in EMAIL_BLOCK_RE.finditer(batch_input_text):
        idx = int(m.group(1))
        raw = m.group(0)
        blocks.append(EmailBlock(idx=idx, raw_block=raw))
    return blocks


def find_latest_batch_run_folder(batch_name: str) -> Path:
    """
    Finds the most recent folder matching batch_name_YYYY...
    """
    candidates = [p for p in PRED_ROOT.glob(f"{batch_name}_*") if p.is_dir()]
    if not candidates:
        raise FileNotFoundError(f"No prediction folders found for {batch_name} in {PRED_ROOT}")
    # folder name contains timestamp; lexicographic sort works with YYYYMMDD_HHMMSS
    candidates.sort(key=lambda p: p.name, reverse=True)
    return candidates[0]


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def normalize(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def url_matches_block(url: str, block_text: str) -> bool:
    if not url:
        return False
    return url.lower() in block_text.lower()


def title_token_score(title: str, block_text: str) -> int:
    """
    Simple lexical overlap score: count meaningful title tokens found in the block.
    """
    t = (title or "").strip()
    if not t:
        return 0
    tokens = [w.lower() for w in re.findall(r"[A-Za-z0-9]+", t) if len(w) >= 4]
    if not tokens:
        return 0
    bt = block_text.lower()
    return sum(1 for tok in set(tokens) if tok in bt)


def infer_event_source_email_index(event: Dict[str, Any], blocks: List[EmailBlock]) -> Tuple[Optional[int], float, str]:
    """
    Try to infer which email block this event came from.
    Returns: (email_idx or None, confidence 0..1, method label)
    Priority:
      1) URL/Registration_URL/Meeting_URL exact inclusion
      2) Title token overlap
    """
    urls = []
    for k in ["URL", "Registration_URL", "Meeting_URL"]:
        v = (event.get(k) or "").strip()
        if v:
            urls.append(v)

    # 1) URL-based linking
    url_hits: List[int] = []
    for b in blocks:
        for u in urls:
            if url_matches_block(u, b.raw_block):
                url_hits.append(b.idx)
                break
    url_hits_unique = sorted(set(url_hits))
    if len(url_hits_unique) == 1:
        return url_hits_unique[0], 0.95, "url_match"
    if len(url_hits_unique) > 1:
        # ambiguous but still useful
        return url_hits_unique[0], 0.60, "url_ambiguous"

    # 2) Title-based linking
    title = event.get("Title") or ""
    scores = [(b.idx, title_token_score(title, b.raw_block)) for b in blocks]
    scores.sort(key=lambda x: x[1], reverse=True)
    best_idx, best_score = scores[0]
    if best_score >= 3:
        return best_idx, 0.75, "title_match_strong"
    if best_score == 2:
        return best_idx, 0.55, "title_match_medium"
    if best_score == 1:
        return best_idx, 0.35, "title_match_weak"

    return None, 0.0, "unlinked"


def group_events_by_email(events: List[Dict[str, Any]], blocks: List[EmailBlock]) -> Dict[int, Dict[str, Any]]:
    """
    Build email-level summaries and link events to email blocks where possible.
    Returns dict keyed by email_idx with:
      - raw_block
      - events (linked)
      - metrics flags
    """
    email_map: Dict[int, Dict[str, Any]] = {
        b.idx: {
            "email_index": b.idx,
            "raw_block": b.raw_block,
            "events": [],
            "link_confidences": [],
        }
        for b in blocks
    }

    for ev in events:
        src_idx, conf, method = infer_event_source_email_index(ev, blocks)
        ev_aug = dict(ev)
        ev_aug["_link_confidence"] = conf
        ev_aug["_link_method"] = method

        if src_idx is not None and src_idx in email_map:
            email_map[src_idx]["events"].append(ev_aug)
            email_map[src_idx]["link_confidences"].append(conf)
        else:
            # keep unlinked events under a special bucket (-1)
            email_map.setdefault(-1, {
                "email_index": -1,
                "raw_block": "",
                "events": [],
                "link_confidences": [],
            })
            email_map[-1]["events"].append(ev_aug)
            email_map[-1]["link_confidences"].append(conf)

    # compute flags
    for idx, rec in email_map.items():
        evs = rec["events"]
        rec["n_events"] = len(evs)
        rec["avg_link_conf"] = (sum(rec["link_confidences"]) / len(rec["link_confidences"])) if rec["link_confidences"] else 0.0

        # flags derived from events
        rec["has_registration"] = any((e.get("Registration_URL") or "").strip() or (e.get("Registration_Needed") is True) for e in evs)
        rec["has_meeting_url"] = any((e.get("Meeting_URL") or "").strip() for e in evs)
        rec["has_location_tba"] = any(normalize(e.get("Location") or "") == "location tba" for e in evs)
        rec["has_sub_event"] = any((e.get("Event_Type") or "").strip() == "sub_event" for e in evs)
        rec["has_main_event"] = any((e.get("Event_Type") or "").strip() == "main_event" for e in evs)
        rec["likely_newsletter"] = rec["n_events"] >= 3 or rec["has_sub_event"]

        # cue-based borderline indicator for negatives (no events but cues exist)
        raw = rec.get("raw_block") or ""
        rec["has_event_cues"] = bool(EVENT_CUE_RE.search(raw))

    return email_map


def pick_one(candidates: List[Dict[str, Any]], used_email_indices: set) -> Optional[Dict[str, Any]]:
    for c in candidates:
        if c["email_index"] not in used_email_indices and c["email_index"] != -1:
            return c
    return None


def select_five_per_batch(email_map: Dict[int, Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Stratified selection (per batch):
      1) one multi-event/newsletter
      2) one registration-heavy
      3) one online/hybrid (meeting link)
      4) one location-risk (TBA)
      5) one borderline/negative (0 events but cues; else 1 event with missing time/date)
    """
    emails = [v for k, v in email_map.items() if k != -1]
    used: set = set()
    chosen: List[Dict[str, Any]] = []

    # 1) multi-event
    multi = sorted([e for e in emails if e["likely_newsletter"]], key=lambda x: (x["n_events"], x["avg_link_conf"]), reverse=True)
    c = pick_one(multi, used)
    if c:
        c["selection_reason"] = "multi_event_or_newsletter"
        chosen.append(c); used.add(c["email_index"])

    # 2) registration-heavy
    reg = sorted([e for e in emails if e["has_registration"] and e["n_events"] > 0], key=lambda x: (x["avg_link_conf"], x["n_events"]), reverse=True)
    c = pick_one(reg, used)
    if c:
        c["selection_reason"] = "registration_required_or_link_present"
        chosen.append(c); used.add(c["email_index"])

    # 3) online/hybrid
    online = sorted([e for e in emails if e["has_meeting_url"] and e["n_events"] > 0], key=lambda x: (x["avg_link_conf"], x["n_events"]), reverse=True)
    c = pick_one(online, used)
    if c:
        c["selection_reason"] = "online_or_hybrid_meeting_link"
        chosen.append(c); used.add(c["email_index"])

    # 4) location-risk
    loc = sorted([e for e in emails if e["has_location_tba"] and e["n_events"] > 0], key=lambda x: (x["avg_link_conf"], x["n_events"]), reverse=True)
    c = pick_one(loc, used)
    if c:
        c["selection_reason"] = "location_tba_or_location_risk"
        chosen.append(c); used.add(c["email_index"])

    # 5) borderline/negative
    negative = sorted([e for e in emails if e["n_events"] == 0 and e["has_event_cues"]], key=lambda x: x["email_index"])
    c = pick_one(negative, used)
    if c:
        c["selection_reason"] = "borderline_negative_event_cues_but_no_events"
        chosen.append(c); used.add(c["email_index"])
    else:
        # fallback: pick a weak extraction (1 event, missing both start date and start time)
        weak = []
        for e in emails:
            if e["email_index"] in used:
                continue
            if e["n_events"] == 1:
                ev = e["events"][0]
                if not ev.get("Start_Date") and not ev.get("Start_Time"):
                    weak.append(e)
        weak.sort(key=lambda x: x["avg_link_conf"])
        c = weak[0] if weak else None
        if c:
            c["selection_reason"] = "borderline_missing_datetime"
            chosen.append(c); used.add(c["email_index"])

    # Fill remaining slots up to 5 with highest-value leftovers (prefer multi-event and higher n_events)
    if len(chosen) < 5:
        leftovers = sorted(
            [e for e in emails if e["email_index"] not in used],
            key=lambda x: (x["likely_newsletter"], x["n_events"], x["avg_link_conf"]),
            reverse=True,
        )
        while len(chosen) < 5 and leftovers:
            c = leftovers.pop(0)
            c["selection_reason"] = "fill_high_value_remaining"
            chosen.append(c); used.add(c["email_index"])

    return chosen[:5]


def save_excel_report(path: Path, rows: List[Dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SelectionReport"

    headers = [
        "global_id",
        "batch",
        "batch_email_index",
        "selection_reason",
        "n_events",
        "likely_newsletter",
        "has_registration",
        "has_meeting_url",
        "has_location_tba",
        "has_sub_event",
        "avg_link_conf",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("global_id"),
            r.get("batch"),
            r.get("email_index"),
            r.get("selection_reason"),
            r.get("n_events"),
            r.get("likely_newsletter"),
            r.get("has_registration"),
            r.get("has_meeting_url"),
            r.get("has_location_tba"),
            r.get("has_sub_event"),
            round(float(r.get("avg_link_conf", 0.0)), 3),
        ])

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--batches", type=str, default="batch_01,batch_02,batch_03,batch_04,batch_05")
    parser.add_argument("--per-batch", type=int, default=5)
    args = parser.parse_args()

    batch_names = [b.strip() for b in args.batches.split(",") if b.strip()]
    if args.per_batch != 5:
        raise ValueError("This script currently implements the fixed 5-per-batch stratified design.")

    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    pack_dir = OUT_ROOT / "pack_for_chatgpt"
    pack_dir.mkdir(parents=True, exist_ok=True)

    selection_rows: List[Dict[str, Any]] = []
    selection_export: Dict[str, Any] = {
        "created_at": datetime.utcnow().isoformat(),
        "strategy": "stratified_5_per_batch_multi_reg_online_location_negative",
        "batches": {},
    }

    global_counter = 0

    for batch_name in batch_names:
        run_dir = find_latest_batch_run_folder(batch_name)
        input_path = run_dir / "input_sent_to_llm.txt"
        events_path = run_dir / "events_filtered.json"

        if not input_path.exists() or not events_path.exists():
            raise FileNotFoundError(f"Missing required files in {run_dir} (need input_sent_to_llm.txt and events_filtered.json)")

        batch_input = input_path.read_text(encoding="utf-8")
        blocks = parse_email_blocks(batch_input)
        events = load_json(events_path)
        if not isinstance(events, list):
            raise ValueError(f"events_filtered.json is not a list in {run_dir}")

        email_map = group_events_by_email(events, blocks)
        chosen = select_five_per_batch(email_map)

        selection_export["batches"][batch_name] = {
            "run_dir": str(run_dir),
            "chosen_email_indices": [c["email_index"] for c in chosen],
        }

        for c in chosen:
            global_counter += 1
            gid = f"email_{global_counter:03d}"
            c["global_id"] = gid
            c["batch"] = batch_name

            # Write the email block text
            email_txt = c["raw_block"]
            (pack_dir / f"{gid}.txt").write_text(email_txt, encoding="utf-8")

            # Write predictions for that email (the linked events)
            # (Include link metadata so later evaluation can see confidence)
            preds = []
            for ev in c["events"]:
                ev_out = {k: ev.get(k) for k in ev.keys() if not k.startswith("_")}
                ev_out["_link_confidence"] = ev.get("_link_confidence")
                ev_out["_link_method"] = ev.get("_link_method")
                preds.append(ev_out)

            (pack_dir / f"{gid}_pred.json").write_text(
                json.dumps(preds, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )

            selection_rows.append(c)

    # Save selection manifest
    (OUT_ROOT / "selection_25.json").write_text(
        json.dumps(selection_export, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Save excel report
    save_excel_report(OUT_ROOT / "selection_report.xlsx", selection_rows)

    print(f"[select_gold_set] Selected {global_counter} emails total.")
    print(f"[select_gold_set] Manifest: {OUT_ROOT / 'selection_25.json'}")
    print(f"[select_gold_set] Pack for ChatGPT: {pack_dir}")
    print(f"[select_gold_set] Report: {OUT_ROOT / 'selection_report.xlsx'}")


if __name__ == "__main__":
    main()
